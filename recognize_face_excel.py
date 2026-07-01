import cv2
import pickle
import numpy as np
from keras_facenet import FaceNet
from datetime import datetime, date
import time
from openpyxl import Workbook, load_workbook
import os
import winsound  # 🔊 ใช้สร้างเสียง beep (Windows เท่านั้น)
import requests  # 📲 ใช้ส่ง LINE Messaging API

# ============================
# 📲 ตั้งค่า LINE Messaging API
# วิธีดู:
#   - CHANNEL_ACCESS_TOKEN: developers.line.biz → Messaging API → Channel access token
#   - USER_ID: developers.line.biz → Basic settings → Your user ID
# ============================
CHANNEL_ACCESS_TOKEN = "/ao0C6kajOjekBAikP634dAjFLXXY/rIjtcCBdJc2Yh7dJq3SDiyRzeaPPGgfzY9a2i/ldgv/1eoGyc+cvXMOUlSecLPs5V1PcqeMvS7k7G5+7m8+LMpY9CuK/Rquoxic/yc9Jra14/XA78h9HHIZAdB04t89/1O/w1cDnyilFU="
USER_ID = "U19c69c5705daeb65714cbdbd90c3ecd4"  # เริ่มด้วย U เช่น Uxxxxxxxx

def send_line_message(message):
    """ส่งข้อความแจ้งเตือนผ่าน LINE Messaging API (Push Message)"""
    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    payload = {
        "to": USER_ID,
        "messages": [
            {
                "type": "text",
                "text": message
            }
        ]
    }
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=5)
        if response.status_code != 200:
            print(f"[LINE Error] {response.status_code}: {response.text}")
    except Exception as e:
        print(f"[LINE Error] {e}")

# ============================
# 🔥 โหลดโมเดล FaceNet
# ใช้แปลง “ใบหน้า → vector (embedding)”
# ============================
embedder = FaceNet()

# ============================
# 🔥 โหลดโมเดล SVM
# ใช้ทำนายว่า embedding นี้เป็น “ใคร”
# ============================
model = pickle.load(open("facenet_svm.pkl", "rb"))

# ============================
# 🔥 โหลด Label Encoder
# ใช้แปลงเลข → ชื่อคน
# ============================
encoder = pickle.load(open("label_encoder.pkl", "rb"))

# ============================
# 🔥 โหลด Haar Cascade
# ใช้ตรวจจับตำแหน่งใบหน้าในภาพ
# ============================
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades +
    "haarcascade_frontalface_default.xml"
)

# ============================
# 🔥 ตั้งค่าไฟล์ Excel สำหรับบันทึกการเข้าเรียน
# ============================
excel_file = "attendance.xlsx"

# ถ้ายังไม่มีไฟล์ → สร้างใหม่
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Date", "Time"])  # header
    wb.save(excel_file)

# โหลดไฟล์ Excel
wb = load_workbook(excel_file)
ws = wb.active

# ============================
# 🔥 ตัวแปรกันเช็คชื่อซ้ำในวันเดียวกัน
# ============================
checked_today = set()

# วันที่ปัจจุบัน
today_date = date.today()

# ============================
# 🔊 ระบบเสียง (กันเสียงรัว)
# ============================
last_beep_time = 0          # เวลาที่ beep ล่าสุด
BEEP_INTERVAL = 3           # เว้น 3 วินาทีค่อย beep อีกครั้ง

# ============================
# 📷 เปิดกล้อง webcam
# ============================
cap = cv2.VideoCapture(0)

while True:

    # อ่านภาพจากกล้อง
    ret, frame = cap.read()
    if not ret:
        break

    # แปลงภาพเป็น grayscale เพื่อใช้ตรวจจับหน้า
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # ============================
    # 🔥 ตรวจจับใบหน้าในภาพ
    # ============================
    faces = face_cascade.detectMultiScale(gray, 1.2, 5)

    for (x, y, w, h) in faces:

        # ============================
        # crop เฉพาะใบหน้า
        # ============================
        face = frame[y:y+h, x:x+w]

        if face.size == 0:
            continue

        # resize ให้ตรงกับ FaceNet (160x160)
        face = cv2.resize(face, (160, 160))

        # แปลง BGR → RGB (FaceNet ใช้ RGB)
        rgb_face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)

        # ============================
        # 🔥 แปลงใบหน้าเป็น embedding vector
        # ============================
        embedding = embedder.embeddings([rgb_face])[0]

        # normalize vector เพื่อให้แม่นขึ้น
        embedding = embedding / np.linalg.norm(embedding)

        # reshape ให้ model ใช้ได้
        embedding = np.expand_dims(embedding, axis=0)

        # ============================
        # 🔥 ทำนายชื่อด้วย SVM
        # ============================
        pred = model.predict(embedding)[0]

        # ความน่าจะเป็นของผลลัพธ์ทั้งหมด
        proba = model.predict_proba(embedding)[0]

        # confidence สูงสุด
        confidence = np.max(proba)

        # แปลงเลข label → ชื่อคน
        name = encoder.inverse_transform([pred])[0]

        # ============================
        # 🔥 ถ้าความมั่นใจต่ำ → Unknown
        # ============================
        if confidence < 0.90:
            name = "Unknown"

        # ============================
        # 🔥 ระบบเช็คชื่อเข้าเรียน
        # ============================
        if name != "Unknown":

            # สร้าง key กันซ้ำ (ชื่อ + วันที่)
            key = f"{name}_{today_date}"

            # ถ้ายังไม่เคยเช็ควันนี้
            if key not in checked_today:

                now = datetime.now()
                time_str = now.strftime("%H:%M:%S")

                # บันทึกลง Excel
                ws.append([name, today_date.strftime("%Y-%m-%d"), time_str])
                wb.save(excel_file)

                # เพิ่มเข้า set กันซ้ำ
                checked_today.add(key)

                print(f"[CHECK-IN] {name} at {time_str}")

                # ============================
                # 📲 ส่งแจ้งเตือน LINE
                # ============================
                send_line_message(
                    f"✅ เช็คชื่อสำเร็จ!\n"
                    f"👤 ชื่อ: {name}\n"
                    f"📅 วันที่: {today_date.strftime('%Y-%m-%d')}\n"
                    f"🕐 เวลา: {time_str}"
                )

                # ============================
                # 🔊 เล่นเสียงแจ้งเตือน
                # ============================
                current_time = time.time()

                # กันเสียงรัว
                if current_time - last_beep_time > BEEP_INTERVAL:
                    winsound.Beep(1200, 200)  # เสียง beep
                    last_beep_time = current_time

        # ============================
        # 🔥 วาดกรอบใบหน้า
        # ============================
        color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)

        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)

        cv2.putText(
            frame,
            f"{name} {confidence:.2f}",
            (x, y-10),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            color,
            2
        )

    # ============================
    # 🔥 แสดงผลหน้าจอ
    # ============================
    cv2.imshow("Attendance System", frame)

    # กด q เพื่อออกจากโปรแกรม
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# ============================
# 🔥 ปิดกล้อง + เซฟไฟล์
# ============================
cap.release()
wb.save(excel_file)
cv2.destroyAllWindows()